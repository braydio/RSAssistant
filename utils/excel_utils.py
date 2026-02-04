"""Excel workbook helpers for account and order management."""

import logging
import os
import shutil
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# Import configuration and functions from init.py
from utils.config_utils import (
    ACCOUNT_MAPPING,
    ERROR_LOG_FILE,
    EXCEL_FILE_MAIN,
    get_account_nickname_or_default,
    load_account_mappings,
    EXCEL_LOGGING_ENABLED,
)
from utils.sql_utils import clear_account_nicknames, sync_account_mappings, upsert_account_mapping

EXCEL_DEPRECATED = True

EXCEL_FILE_DIRECTORY = EXCEL_FILE_MAIN.parent
EXCEL_FILE_NAME = EXCEL_FILE_MAIN.stem
BASE_EXCEL_FILE = EXCEL_FILE_MAIN.name

logger = logging.getLogger(__name__)

# Load Excel log settings
stock_row = 1
date_row = 1
ratio_row = 2
order_row = 3
account_start_row = 4
account_start_column = 1
days_keep_backup = 2


def _get_backup_dates(now=None):
    current = now or datetime.now()
    today_str = current.strftime("%m-%d")
    tomorrow_str = (current + timedelta(days=1)).strftime("%m-%d")
    return today_str, tomorrow_str


def get_excel_file_path(directory=EXCEL_FILE_DIRECTORY, filename=EXCEL_FILE_NAME):
    """Return the path of the base Excel file.

    When :data:`EXCEL_LOGGING_ENABLED` is ``False`` this function simply
    returns the base path without creating or verifying backups.
    """

    base_dir = Path(directory)
    base_excel_file_path = base_dir / BASE_EXCEL_FILE

    if EXCEL_DEPRECATED:
        logger.warning("Excel logging is deprecated; skipping Excel setup.")
        return os.fspath(base_excel_file_path)

    if not EXCEL_LOGGING_ENABLED:
        return os.fspath(base_excel_file_path)

    logger.debug(f"directory={directory}, filename={filename}")

    archive_dir = base_dir / "archive"
    logger.debug(f"archive_dir={archive_dir}")

    today_str, tomorrow_str = _get_backup_dates()
    today_excel_file = archive_dir / f"Backup_{filename}.{today_str}.xlsx"
    tomorrow_excel_file = archive_dir / f"Backup_{filename}.{tomorrow_str}.xlsx"

    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)
        logger.debug(f"Created archive directory: {archive_dir}")

    if not os.path.exists(today_excel_file):
        if os.path.exists(base_excel_file_path):
            shutil.copy(base_excel_file_path, today_excel_file)
            logger.info(f"Created today's backup Excel file: {today_excel_file}")
        else:
            logger.error(f"Base Excel file {base_excel_file_path} not found.")

    if not os.path.exists(tomorrow_excel_file):
        if os.path.exists(today_excel_file):
            shutil.copy(today_excel_file, tomorrow_excel_file)
            logger.info(f"Created tomorrow's backup Excel file: {tomorrow_excel_file}")
        else:
            logger.error(f"Today's backup file {today_excel_file} not found.")

    return os.fspath(base_excel_file_path)


EXCEL_FILE_PATH = get_excel_file_path()
# Example usage


def copy_cell_format(source_cell, target_cell):
    """Copy cell formatting from source to target cell."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format


def create_excel_backups(excel_backup):
    """Create a dated backup of the Excel log when enabled."""

    if EXCEL_DEPRECATED:
        logger.warning("Excel logging is deprecated; skipping backup creation.")
        return

    if not EXCEL_LOGGING_ENABLED:
        logger.info("Excel logging disabled; skipping backup creation.")
        return

    if not os.path.exists(excel_backup):
        shutil.copy(BASE_EXCEL_FILE, excel_backup)
        logger.info(
            f"{BASE_EXCEL_FILE} - New backup created as {excel_backup} from base file."
        )
    else:
        logger.info(f"Active backup file at {excel_backup}")


def excel_backups_checks():
    """Ensure today's and tomorrow's backups exist when enabled."""

    if EXCEL_DEPRECATED:
        return

    if not EXCEL_LOGGING_ENABLED:
        return

    archive_dir = os.path.join(EXCEL_FILE_DIRECTORY, "archive")
    prior_backup = os.path.join(archive_dir, f"Backup_{EXCEL_FILE_NAME}.{today}.xlsx")
    new_backup = os.path.join(archive_dir, f"Backup_{EXCEL_FILE_NAME}.{tomorrow}.xlsx")

    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir, exist_ok=True)
        logger.info(f"Created archive at {archive_dir}")
    if not os.path.exists(prior_backup):
        create_excel_backups(prior_backup)
    if not os.path.exists(new_backup):
        create_excel_backups(new_backup)


def load_excel_workbook(file_path):
    """Load an Excel workbook or return ``None`` if not found.

    Backup checks are skipped when Excel logging is disabled.
    """

    if EXCEL_DEPRECATED:
        logger.warning("Excel logging is deprecated; skipping workbook load.")
        return None

    if not os.path.exists(file_path):
        logger.error(f"Workbook not found: {file_path}")
        return None

    if EXCEL_LOGGING_ENABLED:
        logger.info(f"Loading workbook: {file_path}")
        excel_backups_checks()
    else:
        logger.info("Excel logging disabled; loading workbook without backups.")

    return openpyxl.load_workbook(file_path)


# -- Update Account Mappings


async def index_account_details(
    ctx, excel_main_path=EXCEL_FILE_PATH, mapping_file=ACCOUNT_MAPPING
):
    """Index account details from Excel into JSON storage.

    Excel-backed mapping is deprecated and will log a warning without making
    changes when Excel is disabled.
    """

    if EXCEL_DEPRECATED:
        await ctx.send("Excel mapping is deprecated; no changes were applied.")
        logger.warning("Excel mapping requested while deprecated; skipping.")
        return

    # Load the Excel workbook and select the 'Account Details' sheet
    try:
        wb = load_excel_workbook(excel_main_path)
        if not wb:
            await ctx.send(f"Workbook not found at `{excel_main_path}`.")
            return
        if "Account Details" not in wb.sheetnames:
            await ctx.send("Sheet 'Account Details' not found.")
            return
        ws = wb["Account Details"]
    except Exception as e:
        await ctx.send(f"Error loading Excel file: {e}")
        return

    # Load the existing account mappings using the function from config_utils
    account_mappings = load_account_mappings()

    # Keep track of changes
    changes = []

    # Iterate through the rows of 'Account Details'
    try:
        for row in range(
            2, ws.max_row + 1
        ):  # Start from row 2 (assuming row 1 has headers)
            broker_name = ws[f"A{row}"].value  # Broker name in column A
            group_number = ws[f"B{row}"].value  # Group number in column B
            account_number_raw = ws[f"C{row}"].value  # Account number in column C
            account_nickname = ws[f"D{row}"].value  # Account nickname in column D

            if account_number_raw is None:
                continue

            account_number = str(account_number_raw).zfill(
                4
            )  # Account number padded to 4 digits

            if not account_nickname:
                account_nickname = generate_account_nickname(
                    broker_name,
                    group_number,
                    account_number,
                    mapping_file=mapping_file,
                )
            if (
                not broker_name
                or not group_number
                or not account_number
                or not account_nickname
            ):
                continue  # Skip rows with missing data

            # Ensure all values are strings
            group_number = str(group_number)

            # Ensure the broker and group number exist in the JSON structure
            if broker_name not in account_mappings:
                account_mappings[broker_name] = {}
                changes.append(f"Added broker: {broker_name}")

            if group_number not in account_mappings[broker_name]:
                account_mappings[broker_name][group_number] = {}
                changes.append(
                    f"Added group: {group_number} under broker {broker_name}"
                )

            # Check if account number is already present and if the nickname has changed
            old_nickname = account_mappings[broker_name][group_number].get(
                account_number
            )
            if old_nickname != account_nickname:
                if old_nickname:
                    changes.append(
                        f"Updated account {account_number} under broker {broker_name}, group {group_number} from '{old_nickname}' to '{account_nickname}'"
                    )
                else:
                    changes.append(
                        f"Added new account {account_number} under broker {broker_name}, group {group_number} with nickname '{account_nickname}'"
                    )

            # Add or update the account details under the broker and group number
            account_mappings[broker_name][group_number][
                account_number
            ] = account_nickname

    except Exception as e:
        await ctx.send(f"Error processing Excel rows: {e}")
        return

    # If there were any changes, send them via ctx.send
    if changes:
        change_message = "\n".join(changes)
        logger.info(change_message)
    else:
        await ctx.send("No changes detected in account mappings.")

    try:
        sync_account_mappings(account_mappings)
        await ctx.send("Updated account mappings saved to SQL.")
    except Exception as e:
        await ctx.send(f"Error saving account mappings to SQL: {e}")


async def map_accounts_in_excel_log(
    ctx, filename=EXCEL_FILE_PATH, mapped_accounts_json=ACCOUNT_MAPPING
):
    """Update the Reverse Split Log sheet with mapped accounts.

    Excel updates are deprecated and will no-op when Excel logging is disabled.
    """

    if EXCEL_DEPRECATED:
        await ctx.send("Excel logging is deprecated; no updates were applied.")
        logger.warning("Excel log update requested while deprecated; skipping.")
        return

    # Load the Excel workbook and the Reverse Split Log sheet
    wb = load_excel_workbook(filename)
    if not wb:
        await ctx.send(f"Workbook not found at `{filename}`.")
        return
    if "Reverse Split Log" not in wb.sheetnames:
        await ctx.send("Sheet 'Reverse Split Log' not found.")
        return
    reverse_split_log = wb["Reverse Split Log"]

    # Load the account mappings from the JSON file
    account_mappings = load_account_mappings()

    try:
        # Step 1: Find rows that contain 'Totals' (case-insensitive) and mark them as protected
        protected_rows = (
            set()
        )  # To store the protected rows (Totals, and the rows above and below)

        # Loop through the entire sheet to find 'Totals'
        for row in range(1, reverse_split_log.max_row + 1):
            for col in range(1, reverse_split_log.max_column + 1):
                cell_value = reverse_split_log.cell(row=row, column=col).value
                if (
                    cell_value
                    and isinstance(cell_value, str)
                    and "totals" in cell_value.lower()
                ):
                    # Add the row with 'Totals' and the rows above and below to the protected set
                    protected_rows.add(row)  # The row with 'Totals'
                    if row > 1:  # Protect the row above if it exists
                        protected_rows.add(row - 1)
                    if (
                        row < reverse_split_log.max_row
                    ):  # Protect the row below if it exists
                        protected_rows.add(row + 1)
                    break  # Stop checking other columns once 'Totals' is found in the row

        # Step 2: Count the total number of accounts in the mappings
        total_accounts = sum(
            len(accounts)
            for broker_groups in account_mappings.values()
            for accounts in broker_groups.values()
        )

        # Insert the required number of new rows above the start row
        reverse_split_log.insert_rows(account_start_row, total_accounts)
        protected_rows = {
            row + total_accounts if row >= account_start_row else row
            for row in protected_rows
        }

        # Build a lookup of existing account rows after insertion
        existing_rows = {}
        for row in range(
            account_start_row + total_accounts, reverse_split_log.max_row + 1
        ):
            if row in protected_rows:
                continue
            account_in_log = reverse_split_log.cell(
                row=row, column=account_start_column
            ).value
            if account_in_log:
                existing_rows[str(account_in_log).strip()] = row

        current_row = account_start_row

        # Step 3: Iterate through brokers, group numbers, and accounts in the mapping
        for broker, broker_groups in account_mappings.items():
            for group_number, accounts in broker_groups.items():
                for account_number, nickname in accounts.items():
                    # Skip copying into protected rows
                    while current_row in protected_rows:
                        current_row += 1  # Move to the next unprotected row

                    # Insert the account name in the new rows
                    reverse_split_log.cell(
                        row=current_row,
                        column=account_start_column,
                        value=f"{broker} {nickname}",
                    )

                    # Look for the matching account in the original rows and transfer log data
                    account_key = f"{broker} {nickname}"
                    source_row = existing_rows.get(account_key)
                    if source_row:
                        account_found = True
                        copy_complete_row(reverse_split_log, source_row, current_row)
                    else:
                        account_found = False

                    if not account_found:
                        # If no matching log data is found, leave the log data empty or set a default
                        for col in range(2, reverse_split_log.max_column + 1):
                            reverse_split_log.cell(
                                row=current_row, column=col
                            ).value = None  # Or set a default value

                    current_row += 1  # Move to the next row for the next account

        # Step 4: Delete the original rows after transferring data, skipping protected rows
        rows_to_delete = (
            set(
                range(account_start_row + total_accounts, reverse_split_log.max_row + 1)
            )
            - protected_rows
        )
        for row in sorted(rows_to_delete, reverse=True):
            reverse_split_log.delete_rows(row)

    except KeyError as e:
        await ctx.send(f"Missing key in account mappings: {e}")
        return
    except Exception as e:
        await ctx.send(f"Error updating Excel sheet: {e}")
        return

    # Save the updated workbook
    try:
        save_workbook(wb, EXCEL_FILE_PATH)
        await ctx.send(f"Updated {filename} with account mappings.")
    except Exception as e:
        await ctx.send(f"Error saving Excel file: {e}")


async def clear_account_mappings(ctx, mapping_file=ACCOUNT_MAPPING):
    """Clear the account mappings from SQL storage and notify the user."""

    try:
        cleared = clear_account_nicknames()
        await ctx.send(f"Account mappings have been cleared. ({cleared} SQL rows)")

    except Exception as e:
        await ctx.send(f"Error clearing account mappings: {e}")


async def add_account_mappings(ctx, brokerage, broker_no, account, nickname):
    try:
        upsert_account_mapping(brokerage, broker_no, account, nickname)
        await ctx.send(
            f"Added mapping: {brokerage} - Broker No: {broker_no}, Account: {account}, Nickname: {nickname}"
        )

    except Exception as e:
        await ctx.send(f"An error occurred: {str(e)}")


def copy_complete_row(worksheet, source_row, target_row):
    """Copy both values and formatting from a source row to a target row in the worksheet."""
    for col in range(1, worksheet.max_column + 1):  # Loop over each column
        source_cell = worksheet.cell(row=source_row, column=col)
        target_cell = worksheet.cell(row=target_row, column=col)

        # Copy the cell value
        target_cell.value = source_cell.value

        # Copy formatting (font, fill, border, alignment, and number format)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format


def generate_account_nickname(
    broker_name, group_number, account_number, mapping_file="account_mapping.json"
):
    """
    Generates a placeholder nickname for an account based on broker, group, and account details.
    The nickname format is 'Account N', where N increments if other accounts exist in the same broker/group.

    Parameters:
        broker_name (str): The name of the broker.
        group_number (str): The group number of the account.
        account_number (str): The account number (should be a string padded to 4 digits).
        mapping_file (str): Legacy parameter retained for compatibility; SQL is authoritative.

    Returns:
        str: The generated nickname.
    """

    account_mappings = load_account_mappings()

    # Ensure structure for broker and group exists
    if broker_name not in account_mappings:
        account_mappings[broker_name] = {}
    if group_number not in account_mappings[broker_name]:
        account_mappings[broker_name][group_number] = {}

    # Get existing accounts in the specified broker and group
    existing_accounts = account_mappings[broker_name][group_number]

    # Determine the next available nickname increment
    base_nickname = "Account"
    nickname_increment = 1
    for acc_nickname in existing_accounts.values():
        if acc_nickname.startswith(base_nickname):
            try:
                # Extract and increment the number after 'Account'
                current_number = int(acc_nickname.split()[-1])
                nickname_increment = max(nickname_increment, current_number + 1)
            except ValueError:
                pass

    # Generate the nickname
    new_nickname = f"{base_nickname} {nickname_increment}"

    # Save the new account to the structure
    account_mappings[broker_name][group_number][account_number] = new_nickname

    upsert_account_mapping(broker_name, group_number, account_number, new_nickname)

    return new_nickname


# -- Watchlist New Stock Functions


async def add_stock_to_excel_log(ctx, ticker, split_date, split_ratio):
    """Add a stock ticker to the Excel log.

    Excel logging is deprecated, so this function logs a warning and exits.
    """
    if EXCEL_DEPRECATED:
        await ctx.send("Excel logging is deprecated; no Excel updates were made.")
        logger.warning("Excel log update requested for %s; skipping.", ticker)
        return

    wb = None
    try:
        # Load the Excel workbook and the 'Reverse Split Log' sheet (no await because it's a sync operation)
        wb = load_excel_workbook(EXCEL_FILE_PATH)

        logger.info("Loaded Excel log workbook")

        if not wb:
            logger.error("Workbook could not be loaded.")
            return

        if "Reverse Split Log" not in wb.sheetnames:
            logger.error("Sheet 'Reverse Split Log' not found in the workbook.")
            return
        ws = wb["Reverse Split Log"]

        # Find the last filled column in the row where stock tickers are listed
        last_filled_column = find_last_filled_column(ws, stock_row)
        logger.info(f"Last filled column: {last_filled_column}")

        # Find the next available columns for the ticker and split ratio
        cost_col = last_filled_column + 1
        proceeds_col = cost_col + 1
        spacer_col = proceeds_col + 1
        logger.info(
            f"Columns Letters: Ticker is {get_column_letter(cost_col)}, Date is {get_column_letter(proceeds_col)}"
        )

        # Copy the previous columns to maintain formatting
        previous_cost_col = last_filled_column
        previous_proceeds_col = last_filled_column + 1
        previous_spacer_col = last_filled_column + 2

        copy_column(ws, previous_cost_col, cost_col)
        copy_column(ws, previous_proceeds_col, proceeds_col)
        copy_column(ws, previous_spacer_col, spacer_col)

        # Set ticker, date, ratio, values in the new columns
        ws.cell(row=stock_row, column=cost_col).value = ticker
        ws.cell(row=date_row, column=proceeds_col).value = split_date
        ws.cell(row=ratio_row, column=cost_col).value = "Split Ratio:"
        ws.cell(row=ratio_row, column=proceeds_col).value = split_ratio
        ws.cell(row=order_row, column=cost_col).value = "Cost"
        ws.cell(row=order_row, column=proceeds_col).value = "Proceeds"

        # Save the workbook and close it (no await)
        save_workbook(wb, EXCEL_FILE_PATH)
        logger.info(
            f"Added {ticker} to Excel log at column {get_column_letter(cost_col)} with split date {split_date}."
        )

        await ctx.send(
            f"Added {ticker} to Excel log at column {get_column_letter(cost_col)} with split date {split_date}."
        )

    except Exception as e:
        logger.error(f"Error adding stock to Excel log: {e}")
    finally:
        if wb:
            wb.close()


def copy_column(worksheet, source_col, target_col):
    # Copy values and basic formatting from one column to another.
    for row in range(1, worksheet.max_row + 1):
        source_cell = worksheet.cell(row=row, column=source_col)
        target_cell = worksheet.cell(row=row, column=target_col)

        # Copy the cell value and basic formatting (font, fill, border, alignment)
        target_cell.value = source_cell.value
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format  # Copy number format


def find_last_filled_column(ws, row):
    """Find the last filled column in a given row."""
    last_filled_column = ws.max_column
    for col in range(3, ws.max_column + 1):  # Assuming first two columns are headers
        if ws.cell(row=row, column=col).value is None:
            last_filled_column = col - 1
            break
    return last_filled_column


# -- Logger Functions and Erroring


def update_excel_log(order_data, order_type=None, filename=BASE_EXCEL_FILE):
    """Update the Excel log with buy or sell orders.

    Excel logging is deprecated, so this function logs a warning and exits.
    """
    if EXCEL_DEPRECATED:
        logger.warning("Excel logging is deprecated; skipping order log update.")
        return

    logger.info("Updating excel log.")
    if isinstance(order_data, dict):
        order_data = [order_data]
    elif not isinstance(order_data, list):
        logger.error("order_data must be a list or dict.")
        return

    logger.debug(f"order_data received: {order_data}")

    wb = load_excel_workbook(EXCEL_FILE_PATH)
    if not wb:
        logger.error("Workbook loading failed.")
        return

    ws = get_or_create_sheet(wb, "Reverse Split Log")
    try:
        order_data = validate_order_data(order_data)
        logger.debug(f"Validated order_data: {order_data}")
    except TypeError as e:
        logger.error(f"Invalid order_data format: {str(e)}")
        return

    for order in order_data:
        if not isinstance(order, dict):
            logger.error(
                f"Invalid order format, expected dict but got {type(order)}: {order}"
            )
            continue  # Skip malformed entries
        order_identifier = "unknown"
        try:
            # Extract order details
            broker_name = order["Broker Name"]
            broker_number = int(order["Broker Number"])
            account_number = order["Account Number"]
            order_type = order_type or order["Order Type"]
            stock = order["Stock"]
            quantity = order["Quantity"]
            price = float(order["Price"])
            date = order["Date"]
            order_identifier = f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}"

            # Log the extracted details
            logger.debug(f"Processing order: {order}")

            account_nickname = get_account_nickname_or_default(
                broker_name, broker_number, account_number
            )
            excel_nickname = f"{broker_name} {account_nickname}"
            logger.info(f"Finding for {excel_nickname}")

            # Locate the account row
            account_row = locate_row_for_lookup(ws, excel_nickname, account_start_row)
            if account_row:
                # Locate the stock column
                stock_col = locate_column_for_lookup(
                    ws, stock_row, stock, account_start_column
                )
                if stock_col:
                    # Adjust column if 'sell' type
                    if order_type.lower() == "sell":
                        stock_col += 1

                    # Update the cell value
                    update_cell_value(ws, account_row, stock_col, price)
                    confirm_update = get_column_letter(stock_col)
                    logger.info(
                        f"Updated log for {broker_name} {account_nickname} at cell {confirm_update}{account_row}"
                    )

                    # save_workbook(wb, EXCEL_FILE_PATH)
                    # logger.info(f"Saved excel log workboodk: {wb} filename: {EXCEL_FILE_PATH}")

                    # Remove error logs on success
                    remove_error_from_log(ERROR_LOG_FILE, order_identifier)
                else:
                    # Record the error if the stock is not found
                    error_message = (
                        f"Stock {stock} not found for account {account_nickname}."
                    )
                    record_error_message(error_message, order_identifier)
            else:
                # Record an error if the account row is not found
                error_message = (
                    f"{broker_name} - {account_nickname} not found in Excel."
                )
                record_error_message(error_message, order_identifier)

        except (KeyError, TypeError, ValueError) as e:
            error_message = f"{type(e).__name__}: {str(e)}"
            record_error_message(error_message, order_identifier)

    # Save the workbook and close it after processing
    save_workbook(wb, EXCEL_FILE_PATH)
    logger.info(f"Saved workbook {EXCEL_FILE_PATH}")
    delete_stale_backups(EXCEL_FILE_DIRECTORY, "archive", days_keep_backup)
    wb.close()


def record_error_message(error_message, order_details, error_log_file=ERROR_LOG_FILE):
    """Log an error message to the specified log file and avoid duplicates."""
    formatted_entry = format_error_entry(error_message, order_details)
    if not check_log_for_entry(error_log_file, order_details):
        append_to_log(error_log_file, formatted_entry)
        log_error_order_details(formatted_entry)
    else:
        logger.info(f"Error already exists in error log as details: {order_details}")


def log_error_order_details(order_details):
    logger.info(
        "Deprecating function called log_error_order_details in excel_utils -- returning"
    )
    return
    """Log detailed order information for manual tracking."""
    # if not check_log_for_entry(ERROR_ORDER_DETAILS_FILE, order_details):
    #     append_to_log(ERROR_ORDER_DETAILS_FILE, order_details + "\n")
    #     logger.info(f"Order details logged to {ERROR_ORDER_DETAILS_FILE}")


def remove_error_from_log(file_path, identifier):
    # Remove a block containing the identifier from the specified log file.
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            lines = file.readlines()

        with open(file_path, "w", encoding="utf-8") as file:
            block_to_skip = False
            for line in lines:
                if "--- Error at" in line.strip():
                    block_to_skip = False

                if f"Order Details: {identifier}" in line.strip():
                    block_to_skip = True
                    logger.info(f"Removing block with identifier: {identifier}")

                if not block_to_skip:
                    file.write(line)
    except FileNotFoundError:
        logger.info(f"{file_path} not found. No need to remove anything.")


def delete_stale_backups(
    directory=EXCEL_FILE_DIRECTORY,
    archive_folder="archive",
    days_to_keep=2,
):
    """Delete backup files older than the specified number of days."""

    if EXCEL_DEPRECATED:
        logger.warning("Excel logging is deprecated; skipping backup cleanup.")
        return

    now = datetime.now()
    cutoff = now - timedelta(days=days_to_keep)

    # Debugging: Print directory and archive_folder values

    # Target the archive directory within the specified directory
    archive_dir = os.path.join(os.fspath(directory), str(archive_folder))

    # Ensure the archive directory exists
    if not os.path.exists(archive_dir):
        logger.warning(f"Archive directory does not exist: {archive_dir}")
        return

    # Iterate through files in the archive directory
    for filename in os.listdir(archive_dir):
        # Match files that follow the "Backup_{filename}.{MM-DD}.xlsx" format
        if filename.startswith("Backup") and filename.endswith(".xlsx"):
            # Extract the date part (MM-DD) from the filename
            try:
                # Assuming the format is "Backup_{filename}.{MM-DD}.xlsx"
                date_part = filename.split(".")[-2]  # Extract the date (MM-DD)
                file_date = datetime.strptime(date_part, "%m-%d")

                # Update the file date to the current year for comparison
                file_date = file_date.replace(year=now.year)

                # Check if the file is older than the cutoff date
                if file_date < cutoff:
                    file_path = os.path.join(archive_dir, str(filename))
                    logger.info(f"Deleting old backup file: {filename}")
                    os.remove(file_path)
                    logger.info(f"Deleted old backup file: {filename}")

            except ValueError:
                logger.error(f"Failed to parse date from filename: {filename}")
                continue


# -- logger helpers
def validate_order_data(order_data):
    if isinstance(order_data, dict):
        # Convert single dictionary to a list for consistency
        return [order_data]
    elif isinstance(order_data, list) and all(
        isinstance(order, dict) for order in order_data
    ):
        # If it's already a list of dictionaries, return as-is
        return order_data
    else:
        # Raise an error if the structure is incorrect
        raise TypeError(
            "Expected order_data to be a dictionary or a list of dictionaries."
        )


def get_or_create_sheet(wb, sheet_name):
    # Get or create the specified sheet in the workbook.
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        logger.info(f"'{sheet_name}' sheet was missing, created a new one.")
        return ws


def locate_row_for_lookup(ws, search_value, start_row):
    # Find the row in Column A that matches the account nickname.
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws[f"A{row}"].value
        cell_value = (
            cell_value.strip() if isinstance(cell_value, str) else str(cell_value or "")
        )
        if cell_value.lower() == search_value.strip().lower():
            return row
    return None


def locate_column_for_lookup(ws, in_row, search_value, start_col):
    # Find the column that matches the lookup name.
    for col in range(start_col, ws.max_column + 1, 2):
        if ws.cell(row=in_row, column=col).value == search_value:
            found = get_column_letter(col)
            logger.info(f"Found {search_value} column {found}")
            return col
    return None


def update_cell_value(ws, row, col, value):
    # Update a cell in the specified row and column with the given value.
    ws.cell(row=row, column=col).value = value


def save_workbook(wb, filename):
    """Persist the workbook if Excel logging is enabled."""

    if EXCEL_DEPRECATED:
        logger.warning("Excel logging is deprecated; skipping workbook save.")
        return

    if not EXCEL_LOGGING_ENABLED:
        logger.info("Excel logging disabled; skipping workbook save.")
        return

    try:
        wb.save(filename)
        logger.info(f"Successfully saved the Excel log: {filename}")
    except Exception as e:
        logger.error(f"An error occurred while saving the Excel log: {str(e)}")


# -- Error handling helpers
def check_log_for_entry(log_file_path, entry):
    """Check if the entry already exists in the log file."""
    try:
        with open(log_file_path, "r", encoding="utf-8") as log_file:
            return entry in log_file.read()
    except FileNotFoundError:
        return False


def append_to_log(log_file_path, message):
    """Append a message to the specified log file."""
    with open(log_file_path, "a", encoding="utf-8") as log_file:
        log_file.write(message)
    logger.info(f"Appended to log: {log_file_path}")


def format_error_entry(error_message, order_details):
    """Format the error entry for consistent logger."""
    return f"--- Error at {datetime.now()} ---\nError Message: {error_message}\nOrder Details: {order_details}\n\n"
