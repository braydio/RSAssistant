import json
import logging
import os
import shutil
from copy import copy
from datetime import datetime, timedelta

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Import configuration and functions from init.py
from utils.init import (ACCOUNT_MAPPING_FILE, BASE_EXCEL_FILE, ERROR_LOG_FILE,
                        ERROR_ORDER_DETAILS_FILE, EXCEL_FILE_DIRECTORY,
                        EXCEL_FILE_MAIN_PATH, EXCEL_FILE_NAME, config,
                        get_account_nickname, load_account_mappings,
                        load_config, setup_logging, today, tomorrow)

config = load_config()
# Load Excel log settings
stock_row = config["excel_settings"]["excel_log_settings"]["stock_row"]
date_row = config["excel_settings"]["excel_log_settings"]["date_row"]
ratio_row = config["excel_settings"]["excel_log_settings"]["ratio_row"]
order_row = config["excel_settings"]["excel_log_settings"]["order_row"]
account_start_row = config["excel_settings"]["excel_log_settings"]["account_start_row"]
account_start_column = config["excel_settings"]["excel_log_settings"]["account_start_column"]
days_keep_backup = config["excel_settings"]["excel_file_settings"]["days_keep_backup"]

setup_logging(config)


def copy_cell_format(source_cell, target_cell):
    """Copy cell formatting from source to target cell."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format


def create_excel_backups(excel_backup):
    if not os.path.exists(excel_backup):
        shutil.copy(BASE_EXCEL_FILE, excel_backup)
        logging.info(
            f"{BASE_EXCEL_FILE} - New backup created as {excel_backup} from base file."
        )
    else:
        logging.info(f"Active backup file at {excel_backup}")


def excel_backups_checks():
    # Save backups, create archive directory
    archive_dir = os.path.join(EXCEL_FILE_DIRECTORY, "archive")
    prior_backup = os.path.join(archive_dir, f"Backup_{EXCEL_FILE_NAME}.{today}.xlsx")
    new_backup = os.path.join(archive_dir, f"Backup_{EXCEL_FILE_NAME}.{tomorrow}.xlsx")

    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir, exist_ok=True)
        logging.info(f"Created archive at {archive_dir}")
    if not os.path.exists(prior_backup):
        create_excel_backups(prior_backup)
    if not os.path.exists(new_backup):
        create_excel_backups(new_backup)


def load_excel_workbook(file_path):
    """Load an Excel workbook or return None if not found."""
    if os.path.exists(file_path):
        logging.info(f"Loading workbook: {file_path}")
        excel_backups_checks()
        return openpyxl.load_workbook(file_path)
    else:
        logging.error(f"Workbook not found: {file_path}")
        return None


# -- Update Account Mappings


async def index_account_details(
    ctx, filename=EXCEL_FILE_MAIN_PATH, mapping_file=ACCOUNT_MAPPING_FILE
):
    """Index account details from an Excel file, update account mappings in JSON, and notify about changes."""

    # Load the Excel workbook and select the 'Account Details' sheet
    try:
        wb = openpyxl.load_workbook(filename)
        if "Account Details" not in wb.sheetnames:
            await ctx.send("Sheet 'Account Details' not found.")
            return
        ws = wb["Account Details"]
    except Exception as e:
        await ctx.send(f"Error loading Excel file: {e}")
        return

    # Load the existing account mappings using the function from config_utils
    account_mappings = load_account_mappings()

    # Keep track of changes
    changes = []

    # Iterate through the rows of 'Account Details'
    try:
        for row in range(
            2, ws.max_row + 1
        ):  # Start from row 2 (assuming row 1 has headers)
            broker_name = ws[f"A{row}"].value  # Broker name in column A
            group_number = ws[f"B{row}"].value  # Group number in column B
            account_number = str(ws[f"C{row}"].value).zfill(
                4
            )  # Account number in column C, padded to 4 digits
            account_nickname = ws[f"D{row}"].value  # Account nickname in column D

            if not account_nickname:
                account_nickname = generate_account_nickname(
                    broker_name,
                    group_number,
                    account_number,
                    mapping_file=ACCOUNT_MAPPING_FILE,
                )
            if (
                not broker_name
                or not group_number
                or not account_number
                or not account_nickname
            ):
                continue  # Skip rows with missing data

            # Ensure all values are strings
            group_number = str(group_number)

            # Ensure the broker and group number exist in the JSON structure
            if broker_name not in account_mappings:
                account_mappings[broker_name] = {}
                changes.append(f"Added broker: {broker_name}")

            if group_number not in account_mappings[broker_name]:
                account_mappings[broker_name][group_number] = {}
                changes.append(
                    f"Added group: {group_number} under broker {broker_name}"
                )

            # Check if account number is already present and if the nickname has changed
            old_nickname = account_mappings[broker_name][group_number].get(
                account_number
            )
            if old_nickname != account_nickname:
                if old_nickname:
                    changes.append(
                        f"Updated account {account_number} under broker {broker_name}, group {group_number} from '{old_nickname}' to '{account_nickname}'"
                    )
                else:
                    changes.append(
                        f"Added new account {account_number} under broker {broker_name}, group {group_number} with nickname '{account_nickname}'"
                    )

            # Add or update the account details under the broker and group number
            account_mappings[broker_name][group_number][
                account_number
            ] = account_nickname

    except Exception as e:
        await ctx.send(f"Error processing Excel rows: {e}")
        return

    # If there were any changes, send them via ctx.send
    if changes:
        change_message = "\n".join(changes)
        logging.info(change_message)
    else:
        await ctx.send("No changes detected in account mappings.")

    # Save the updated mappings back to the JSON file
    try:
        with open(mapping_file, "w") as f:
            json.dump(account_mappings, f, indent=4)
        await ctx.send(f"Updated mappings saved to `{mapping_file}`.")

    except Exception as e:
        await ctx.send(f"Error saving JSON file: {e}")


async def map_accounts_in_excel_log(
    ctx, filename=EXCEL_FILE_MAIN_PATH, mapped_accounts_json=ACCOUNT_MAPPING_FILE
):
    """Update the Reverse Split Log sheet by inserting new rows, copying data and formatting, and deleting original rows."""

    # Load the Excel workbook and the Reverse Split Log sheet
    wb = openpyxl.load_workbook(filename)
    reverse_split_log = wb["Reverse Split Log"]

    # Load the account mappings from the JSON file
    with open(mapped_accounts_json, "r") as f:
        account_mappings = json.load(f)

    try:
        # Step 1: Find rows that contain 'Totals' (case-insensitive) and mark them as protected
        protected_rows = (
            set()
        )  # To store the protected rows (Totals, and the rows above and below)

        # Loop through the entire sheet to find 'Totals'
        for row in range(1, reverse_split_log.max_row + 1):
            for col in range(1, reverse_split_log.max_column + 1):
                cell_value = reverse_split_log.cell(row=row, column=col).value
                if (
                    cell_value
                    and isinstance(cell_value, str)
                    and "totals" in cell_value.lower()
                ):
                    # Add the row with 'Totals' and the rows above and below to the protected set
                    protected_rows.add(row)  # The row with 'Totals'
                    if row > 1:  # Protect the row above if it exists
                        protected_rows.add(row - 1)
                    if (
                        row < reverse_split_log.max_row
                    ):  # Protect the row below if it exists
                        protected_rows.add(row + 1)
                    break  # Stop checking other columns once 'Totals' is found in the row

        # Step 2: Count the total number of accounts in the mappings
        total_accounts = sum(
            len(accounts)
            for broker_groups in account_mappings.values()
            for accounts in broker_groups.values()
        )

        # Insert the required number of new rows above the start row
        reverse_split_log.insert_rows(account_start_row, total_accounts)

        current_row = account_start_row

        # Step 3: Iterate through brokers, group numbers, and accounts in the mapping
        for broker, broker_groups in account_mappings.items():
            for group_number, accounts in broker_groups.items():
                for account_number, nickname in accounts.items():
                    # Skip copying into protected rows
                    while current_row in protected_rows:
                        current_row += 1  # Move to the next unprotected row

                    # Insert the account name in the new rows
                    reverse_split_log.cell(
                        row=current_row,
                        column=account_start_column,
                        value=f"{broker} {nickname}",
                    )

                    # Look for the matching account in the original rows and transfer log data
                    account_found = False
                    for row in range(
                        account_start_row + total_accounts,
                        reverse_split_log.max_row + 1,
                    ):
                        account_in_log = reverse_split_log.cell(
                            row=row, column=account_start_column
                        ).value
                        if account_in_log == f"{broker} {nickname}":
                            # Ensure we don't overwrite protected rows during copying
                            if row not in protected_rows:
                                account_found = True
                                # Copy both values and formatting from the original row to the new row
                                copy_complete_row(reverse_split_log, row, current_row)
                            break

                    if not account_found:
                        # If no matching log data is found, leave the log data empty or set a default
                        for col in range(2, reverse_split_log.max_column + 1):
                            reverse_split_log.cell(
                                row=current_row, column=col
                            ).value = None  # Or set a default value

                    current_row += 1  # Move to the next row for the next account

        # Step 4: Delete the original rows after transferring data, skipping protected rows
        rows_to_delete = (
            set(
                range(account_start_row + total_accounts, reverse_split_log.max_row + 1)
            )
            - protected_rows
        )
        for row in sorted(rows_to_delete, reverse=True):
            reverse_split_log.delete_rows(row)

    except KeyError as e:
        await ctx.send(f"Missing key in account mappings: {e}")
        return
    except Exception as e:
        await ctx.send(f"Error updating Excel sheet: {e}")
        return

    # Save the updated workbook
    try:
        wb.save(filename)
        await ctx.send(f"Updated {filename} with account mappings.")
    except Exception as e:
        await ctx.send(f"Error saving Excel file: {e}")


async def clear_account_mappings(ctx, mapping_file=ACCOUNT_MAPPING_FILE):
    """Clear the account mappings JSON file and notify the user."""

    try:
        # Clear the account mappings by writing an empty dictionary to the file
        with open(mapping_file, "w") as f:
            json.dump({}, f, indent=4)

        # Notify the user that the file has been cleared
        await ctx.send(f"Account mappings in `{mapping_file}` have been cleared.")

    except Exception as e:
        await ctx.send(f"Error clearing the JSON file: {e}")


def copy_complete_row(worksheet, source_row, target_row):
    """Copy both values and formatting from a source row to a target row in the worksheet."""
    for col in range(1, worksheet.max_column + 1):  # Loop over each column
        source_cell = worksheet.cell(row=source_row, column=col)
        target_cell = worksheet.cell(row=target_row, column=col)

        # Copy the cell value
        target_cell.value = source_cell.value

        # Copy formatting (font, fill, border, alignment, and number format)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format


def generate_account_nickname(
    broker_name, group_number, account_number, mapping_file="account_mapping.json"
):
    """
    Generates a placeholder nickname for an account based on broker, group, and account details.
    The nickname format is 'Account N', where N increments if other accounts exist in the same broker/group.

    Parameters:
        broker_name (str): The name of the broker.
        group_number (str): The group number of the account.
        account_number (str): The account number (should be a string padded to 4 digits).
        mapping_file (str): Path to the JSON mapping file.

    Returns:
        str: The generated nickname.
    """

    # Load current account mappings from JSON file
    try:
        with open(ACCOUNT_MAPPING_FILE, "r") as f:
            account_mappings = json.load(f)
    except FileNotFoundError:
        logging.info(f"")
        account_mappings = {}  # Initialize empty if file doesn't exist

    # Ensure structure for broker and group exists
    if broker_name not in account_mappings:
        account_mappings[broker_name] = {}
    if group_number not in account_mappings[broker_name]:
        account_mappings[broker_name][group_number] = {}

    # Get existing accounts in the specified broker and group
    existing_accounts = account_mappings[broker_name][group_number]

    # Determine the next available nickname increment
    base_nickname = "Account"
    nickname_increment = 1
    for acc_nickname in existing_accounts.values():
        if acc_nickname.startswith(base_nickname):
            try:
                # Extract and increment the number after 'Account'
                current_number = int(acc_nickname.split()[-1])
                nickname_increment = max(nickname_increment, current_number + 1)
            except ValueError:
                pass

    # Generate the nickname
    new_nickname = f"{base_nickname} {nickname_increment}"

    # Save the new account to the structure
    account_mappings[broker_name][group_number][account_number] = new_nickname

    # Write updated mappings back to JSON
    with open(mapping_file, "w") as f:
        json.dump(account_mappings, f, indent=4)

    return new_nickname


# -- Watchlist New Stock Functions


async def add_stock_to_excel_log(ctx, ticker, split_date, split_ratio):
    """Add the given stock ticker to the next available spot in the Excel log and copy formatting from the previous columns."""
    try:
        # Load the Excel workbook and the 'Reverse Split Log' sheet (no await because it's a sync operation)
        wb = EXCEL_FILE_MAIN_PATH
        load_excel_log(wb)
        logging.info("Loaded Excel log workbook")

        if not wb:
            logging.error("Workbook could not be loaded.")
            return

        if "Reverse Split Log" not in wb.sheetnames:
            logging.error("Sheet 'Reverse Split Log' not found in the workbook.")
            return
        ws = wb["Reverse Split Log"]

        # Find the last filled column in the row where stock tickers are listed
        last_filled_column = find_last_filled_column(ws, stock_row)
        logging.info(f"Last filled column: {last_filled_column}")

        # Find the next available columns for the ticker and split ratio
        cost_col = last_filled_column + 1
        proceeds_col = cost_col + 1
        spacer_col = proceeds_col + 1
        logging.info(
            f"Columns Letters: Ticker is {get_column_letter(cost_col)}, Date is {get_column_letter(proceeds_col)}"
        )

        # Copy the previous columns to maintain formatting
        copy_column(ws, cost_col, spacer_col)
        copy_column(ws, cost_col, proceeds_col)
        copy_column(ws, cost_col, spacer_col)

        # Set ticker, date, ratio, values in the new columns
        ws.cell(row=stock_row, column=cost_col).value = ticker
        ws.cell(row=date_row, column=proceeds_col).value = split_date
        ws.cell(row=ratio_row, column=cost_col).value = "Split Ratio:"
        ws.cell(row=ratio_row, column=proceeds_col).value = split_ratio
        ws.cell(row=order_row, column=cost_col).value = "Cost"
        ws.cell(row=order_row, column=proceeds_col).value = "Proceeds"

        # Save the workbook and close it (no await)
        wb.save(BASE_EXCEL_FILE)
        logging.info(
            f"Added {ticker} to Excel log at column {get_column_letter(cost_col)} with split date {split_date}."
        )

        await ctx.send(
            f"Added {ticker} to Excel log at column {get_column_letter(cost_col)} with split date {split_date}."
        )

    except Exception as e:
        logging.error(f"Error adding stock to Excel log: {e}")
    finally:
        if wb:
            wb.close()


def copy_column(worksheet, source_col, target_col):
    # Copy values and basic formatting from one column to another.
    for row in range(1, worksheet.max_row + 1):
        source_cell = worksheet.cell(row=row, column=source_col)
        target_cell = worksheet.cell(row=row, column=target_col)

        # Copy the cell value and basic formatting (font, fill, border, alignment)
        target_cell.value = source_cell.value
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format  # Copy number format


def find_last_filled_column(ws, row):
    """Find the last filled column in a given row."""
    last_filled_column = ws.max_column
    for col in range(3, ws.max_column + 1):  # Assuming first two columns are headers
        if ws.cell(row=row, column=col).value is None:
            last_filled_column = col - 1
            break
    return last_filled_column


# -- Logger Functions and Erroring


def update_excel_log(
    order_data, order_type=None, filename=BASE_EXCEL_FILE, config=config
):
    """Update the Excel log with buy or sell orders, handling single or multiple orders."""
    logging.info("Updating excel log.")
    if isinstance(order_data, dict):
        order_data = [order_data]
    elif not isinstance(order_data, list):
        logging.error("order_data must be a list or dict.")
        return
    
    logging.debug(f"order_data received: {order_data}")

    wb = load_excel_workbook(EXCEL_FILE_MAIN_PATH)
    if not wb:
        logging.error("Workbook loading failed.")
        return

    ws = get_or_create_sheet(wb, "Reverse Split Log")
    try:
        order_data = validate_order_data(order_data)
        logging.debug(f"Validated order_data: {order_data}")
    except TypeError as e:
        logging.error(f"Invalid order_data format: {str(e)}")
        return

    for order in order_data:
        if not isinstance(order, dict):
            logging.error(f"Invalid order format, expected dict but got {type(order)}: {order}")
            continue  # Skip malformed entries
        try:
            # Extract order details
            broker_name = order["Broker Name"]
            broker_number = int(order["Broker Number"])
            account_number = order["Account Number"]
            order_type = order_type or order["Order Type"]
            stock = order["Stock"]
            quantity = order["Quantity"]
            price = float(order["Price"])
            date = order["Date"]

            # Log the extracted details
            logging.debug(f"Processing order: {order}")

            account_nickname = get_account_nickname(
                broker_name, broker_number, account_number
            )
            excel_nickname = f"{broker_name} {account_nickname}"
            logging.info(f"Finding for {excel_nickname}")

            # Locate the account row
            account_row = locate_row_for_lookup(ws, excel_nickname, account_start_row)
            if account_row:
                # Locate the stock column
                stock_col = locate_column_for_lookup(
                    ws, stock_row, stock, account_start_column
                )
                if stock_col:
                    # Adjust column if 'sell' type
                    if order_type.lower() == "sell":
                        stock_col += 1

                    # Update the cell value
                    update_cell_value(ws, account_row, stock_col, price)
                    confirm_update = get_column_letter(stock_col)
                    logging.info(
                        f"Updated log for {broker_name} {account_nickname} at cell {confirm_update}{account_row}"
                    )

                    # Remove error logs on success
                    identifier = f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}"
                    remove_error_from_log(ERROR_LOG_FILE, identifier)
                    remove_error_from_log(ERROR_ORDER_DETAILS_FILE, identifier)
                else:
                    # Record the error if the stock is not found
                    error_message = (
                        f"Stock {stock} not found for account {account_nickname}."
                    )
                    record_error_message(
                        error_message,
                        f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}",
                    )
            else:
                # Record an error if the account row is not found
                error_message = (
                    f"{broker_name} - {account_nickname} not found in Excel."
                )
                record_error_message(
                    error_message,
                    f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}",
                )

        except ValueError as e:
            # Log a ValueError specifically
            error_message = f"ValueError: {str(e)}"
            record_error_message(
                error_message,
                f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}",
            )

    # Save the workbook and close it after processing
    save_workbook(wb, filename)
    delete_stale_backups(EXCEL_FILE_DIRECTORY, "archive", days_keep_backup)
    wb.close()


def record_error_message(error_message, order_details, error_log_file=ERROR_LOG_FILE):
    """Log an error message to the specified log file and avoid duplicates."""
    formatted_entry = format_error_entry(error_message, order_details)
    if not check_log_for_entry(error_log_file, order_details):
        append_to_log(error_log_file, formatted_entry)
        log_error_order_details(formatted_entry)
    else:
        logging.info(f"Error already exists in error log as details: {order_details}")


def log_error_order_details(order_details):
    """Log detailed order information for manual tracking."""    
    if not check_log_for_entry(ERROR_ORDER_DETAILS_FILE, order_details):
        append_to_log(ERROR_ORDER_DETAILS_FILE, order_details + "\n")
        logging.info(f"Order details logged to {ERROR_ORDER_DETAILS_FILE}")


def remove_error_from_log(file_path, identifier):
    # Remove a block containing the identifier from the specified log file.
    try:
        with open(file_path, "r") as file:
            lines = file.readlines()

        with open(file_path, "w") as file:
            block_to_skip = False
            for line in lines:
                if "--- Error at" in line.strip():
                    block_to_skip = False

                if f"Order Details: {identifier}" in line.strip():
                    block_to_skip = True
                    logging.info(f"Removing block with identifier: {identifier}")

                if not block_to_skip:
                    file.write(line)
    except FileNotFoundError:
        logging.info(f"{file_path} not found. No need to remove anything.")


def delete_stale_backups(
    directory=EXCEL_FILE_DIRECTORY,
    archive_folder="archive",
    days_to_keep=days_keep_backup,
):
    """Delete backup files older than the specified number of days."""
    now = datetime.now()
    cutoff = now - timedelta(days=days_to_keep)

    # Debugging: Print directory and archive_folder values

    # Target the archive directory within the specified directory
    archive_dir = os.path.join(str(directory), str(archive_folder))

    # Ensure the archive directory exists
    if not os.path.exists(archive_dir):
        logging.warning(f"Archive directory does not exist: {archive_dir}")
        return

    # Target the archive directory within the specified directory
    archive_dir = os.path.join(str(directory), str(archive_folder))

    # Ensure the archive directory exists
    if not os.path.exists(archive_dir):
        logging.warning(f"Archive directory does not exist: {archive_dir}")
        return

    # Iterate through files in the archive directory
    for filename in os.listdir(archive_dir):
        # Match files that follow the "Backup_{filename}.{MM-DD}.xlsx" format
        if filename.startswith("Backup") and filename.endswith(".xlsx"):
            # Extract the date part (MM-DD) from the filename
            try:
                # Assuming the format is "Backup_{filename}.{MM-DD}.xlsx"
                date_part = filename.split(".")[-2]  # Extract the date (MM-DD)
                file_date = datetime.strptime(date_part, "%m-%d")

                # Update the file date to the current year for comparison
                file_date = file_date.replace(year=now.year)

                # Check if the file is older than the cutoff date
                if file_date < cutoff:
                    file_path = os.path.join(archive_dir, str(filename))
                    logging.info(f"Deleting old backup file: {filename}")
                    os.remove(file_path)
                    logging.info(f"Deleted old backup file: {filename}")

            except ValueError:
                logging.error(f"Failed to parse date from filename: {filename}")
                continue


# -- Logging helpers
def validate_order_data(order_data):
    if isinstance(order_data, dict):
        # Convert single dictionary to a list for consistency
        return [order_data]
    elif isinstance(order_data, list) and all(
        isinstance(order, dict) for order in order_data
    ):
        # If it's already a list of dictionaries, return as-is
        return order_data
    else:
        # Raise an error if the structure is incorrect
        raise TypeError(
            "Expected order_data to be a dictionary or a list of dictionaries."
        )


def get_or_create_sheet(wb, sheet_name):
    # Get or create the specified sheet in the workbook.
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        logging.info(f"'{sheet_name}' sheet was missing, created a new one.")
        return ws


def locate_row_for_lookup(ws, search_value, start_row):
    # Find the row in Column A that matches the account nickname.
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws[f"A{row}"].value
        cell_value = (
            cell_value.strip() if isinstance(cell_value, str) else str(cell_value or "")
        )
        if cell_value.lower() == search_value.strip().lower():
            return row
    return None


def locate_column_for_lookup(ws, in_row, search_value, start_col):
    # Find the column that matches the lookup name.
    for col in range(start_col, ws.max_column + 1, 2):
        if ws.cell(row=in_row, column=col).value == search_value:
            found = get_column_letter(col)
            logging.info(f"Found {search_value} column {found}")
            return col
    return None


def update_cell_value(ws, row, col, value):
    # Update a cell in the specified row and column with the given value.
    ws.cell(row=row, column=col).value = value


def save_workbook(wb, filename):
    # Save the workbook and handle any errors.
    try:
        wb.save(filename)
        logging.info(f"Successfully saved the Excel log: {filename}")
    except Exception as e:
        logging.error(f"An error occurred while saving the Excel log: {str(e)}")


# -- Error handling helpers
def check_log_for_entry(log_file_path, entry):
    """Check if the entry already exists in the log file."""
    try:
        with open(log_file_path, "r") as log_file:
            return entry in log_file.read()
    except FileNotFoundError:
        return False


def append_to_log(log_file_path, message):
    """Append a message to the specified log file."""
    with open(log_file_path, "a") as log_file:
        log_file.write(message)
    logging.info(f"Appended to log: {log_file_path}")


def format_error_entry(error_message, order_details):
    """Format the error entry for consistent logging."""
    return f"--- Error at {datetime.now()} ---\nError Message: {error_message}\nOrder Details: {order_details}\n\n"


# -- OLD STUFF TO REMOVE

"""
def log_error_message(error_message, order_details, error_log_file=ERROR_LOG_FILE):
    # Log an error message to the specified log file and avoid duplicate entries.
    logging.info(f"Logging error message, order details: {order_details}")
    try:
        with open(error_log_file, 'r') as log_file:
            log_contents = log_file.read()
    except FileNotFoundError:
        log_contents = ""  # If the log file doesn't exist, we'll create it later

    if order_details in log_contents:
        logging.info(f"Order details already logged as error: {order_details}")
        return  # Avoid logging duplicates

    # Append the error message and order details to the log file
    with open(error_log_file, 'a') as log_file:
        log_file.write(f"--- Error at {datetime.now()} ---\n")
        log_file.write(f"Error Message: {error_message}\n")
        log_file.write(f"Order Details: {order_details}\n\n")
    
    logging.info(f"Written to log file: {error_log_file}")
    log_error_order_details(order_details)

def log_error_order_details(order_details):
    logging.info(order_details)  # Ensure this is not outputting the raw dictionary with colons
    
    # Log the order details for later manual entry for errors in a separate file.
    try:
        logging.info(order_details)
        with open(ERROR_ORDER_DETAILS_FILE, 'r') as order_file:
            existing_orders = order_file.read()
    except FileNotFoundError:
        existing_orders = ""

    # Format order_details as a clean string to avoid colons
    formatted_order = (
        f"Broker: {order_details['broker_name']}, "
        f"Group Number: {order_details['group_number']}, "
        f"Account: {order_details['account']}, "
        f"Order Type: {order_details['order_type']}, "
        f"Stock: {order_details['stock']}, "
        f"Price: {order_details['price']}"
    )

    # Avoid logging duplicate orders
    if formatted_order not in existing_orders:
        with open(ERROR_ORDER_DETAILS_FILE, 'a') as order_file:
            order_file.write(formatted_order + '\n')
        logging.info(f"Order details saved to {ERROR_ORDER_DETAILS_FILE}")

    # Log formatted order for consistency
    logging.info(formatted_order)

def remove_from_file(file_path, identifier):
    Remove a specific block containing the identifier from the file.
    logging.info(f'Removing {identifier} from {file_path}')
    try:
        with open(file_path, 'r') as file:
            lines = file.readlines()

        with open(file_path, 'w') as file:
            block_to_skip = False
            for line in lines:
                stripped_line = line.strip()

                if "--- Error at" in stripped_line:
                    block_to_skip = False

                if f"Order Details: {identifier}" in stripped_line:
                    block_to_skip = True
                    logging.info(f"Skipping block with identifier: {identifier}")

                if not block_to_skip:
                    file.write(line)

    except FileNotFoundError:
        logging.info(f"{file_path} not found.")

def update_excel_log(order_data, order_type=None, filename=excel_log_file, config=config):
    
    wb = None
    error_log_file = ERROR_LOG_FILE

    # Convert order_data to a list if it's a single dictionary
    if isinstance(order_data, dict):
        order_data = [order_data]

    try:
        # Load the Excel workbook
        wb = load_workbook(filename)
        if not wb:
            logging.error(f"Workbook could not be loaded: {filename}")
            return
        
        # Validate and normalize order_data
        try:
            order_data = validate_order_data(order_data)
        except TypeError as e:
            logging.error(f"Invalid order_data format: {str(e)}")
            return  # Exit function if validation fails

        # Access or create the "Reverse Split Log" sheet
        if "Reverse Split Log" in wb.sheetnames:
            ws = wb["Reverse Split Log"]
        else:
            ws = wb.create_sheet("Reverse Split Log")
            logging.info(f"'Reverse Split Log' sheet was missing, created a new one.")

        logging.info(order_data)  # Debugging: Print the orders to verify they are correct

        for order in order_data:
            try:
                # Extract details and get the account nickname
                broker_name = order['Broker Name']
                broker_number = int(order['Broker Number'])
                account_number = order['Account Number']
                order_type = order_type or order['Order Type']
                stock = order['Stock']
                quantity = order['Quantity']
                price = float(order['Price'])  # Convert np.float64 to regular float
                date = order['Date']

                error_order = f"manual {broker_name} {broker_number} {account_number} {order_type} {stock} {price}"

                account_nickname = get_account_nickname(broker_name, broker_number, account_number)
                excel_nickname = f"{broker_name} {account_nickname}"
                logging.info(f"Processing {order_type} order for {excel_nickname}, stock: {stock}, price: {price}")

                # Locate the row for the account in Column A based on account_nickname
                account_row = None
                for row in range(account_start_row, ws.max_row + 1):
                    cell_value = ws[f'A{row}'].value
                    cell_value = cell_value.strip() if isinstance(cell_value, str) else str(cell_value or '')

                    if cell_value.lower() == excel_nickname.strip().lower():
                        account_row = row
                        break

                if account_row:
                    # Locate the stock column in the specified stock row
                    stock_col = None
                    for col in range(3, ws.max_column + 1, 2):
                        if ws.cell(row=stock_row, column=col).value == stock:
                            stock_col = col + 1 if order_type.lower() == 'sell' else col
                            break

                    if stock_col:
                        ws.cell(row=account_row, column=stock_col).value = price
                        logging.info(f"{excel_nickname}: Updated column {stock_col}, row {account_row} with {price} for {order_type} on {stock}.")
                        # Remove the corresponding error from the error logs if the order was successfully processed
                        remove_from_file(error_log_file, error_order)
                        remove_from_file(ERROR_ORDER_DETAILS_FILE, error_order)
                    else:
                        error_message = f"Stock {stock} not found for account {account_nickname}."
                        log_error_message(error_message, f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}", error_log_file)
                else:
                    error_message = f"{broker_name} - {account_nickname} not found in Excel."
                    log_error_message(error_message, f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}", error_log_file)

            except ValueError as e:
                error_message = f"ValueError: {str(e)}"
                log_error_message(error_message, f"{broker_name} {broker_number} {account_number} {order_type} {stock} {price}", error_log_file)

        try:
            wb.save(filename)
            logging.info(f"Saved Excel file: {filename}")
            logging.info(f"Successfully saved the Excel log: {filename}")

            # Manage backups by removing stale ones (using the config value)
            delete_stale_backups(EXCEL_FILE_DIRECTORY, 'archive', days_keep_backup)
        

        except Exception as e:
            logging.error(f"An error occurred while updating the Excel log: {str(e)}")

    finally:
        if wb:
            wb.close()

"""